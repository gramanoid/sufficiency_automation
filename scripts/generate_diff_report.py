#!/usr/bin/env python3
"""
Diff Report Generator - Compares PPT data against Excel data
Generates comprehensive report of all discrepancies
PPT is source of truth - Excel needs to be updated to match
"""

import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Comparison tolerances
TOLERANCE_CURRENCY = 1.0  # £1 tolerance for rounding
TOLERANCE_PERCENTAGE = 0.001  # 0.1% tolerance
TOLERANCE_INTEGER = 0  # Exact match for campaign counts

# Fields to compare
COMPARE_FIELDS = {
    'budget_2026': {'type': 'currency', 'label': '2026 Budget'},
    'sufficient_2026': {'type': 'currency', 'label': '2026 Sufficient'},
    'gap_gbp': {'type': 'currency', 'label': 'Gap (GBP)'},
    'gap_pct': {'type': 'percentage', 'label': 'Gap (%)'},
    'awa': {'type': 'percentage', 'label': 'AWA'},
    'con': {'type': 'percentage', 'label': 'CON'},
    'pur': {'type': 'percentage', 'label': 'PUR'},
    'tv': {'type': 'percentage', 'label': 'TV'},
    'digital': {'type': 'percentage', 'label': 'Digital'},
    'others': {'type': 'percentage', 'label': 'Others'},
    'long_campaigns': {'type': 'integer', 'label': 'Long Campaigns'},
    'short_campaigns': {'type': 'integer', 'label': 'Short Campaigns'},
    'long_pct': {'type': 'percentage', 'label': 'Long %'},
}

def normalize_brand(brand):
    """Normalize brand name for matching"""
    if not brand:
        return ''
    return brand.strip().upper().replace(' ', '').replace('-', '')

def normalize_market(market):
    """Normalize market name"""
    if not market:
        return ''
    return market.strip().upper()

def make_key(record):
    """Create unique key for matching records"""
    market = normalize_market(record.get('market', ''))
    category = (record.get('category') or '').strip().upper()
    brand = normalize_brand(record.get('brand', ''))
    return (market, category, brand)

def compare_values(ppt_val, excel_val, field_type):
    """Compare two values and return (match, diff_amount, diff_pct)"""
    # Handle None/missing values
    if ppt_val is None and excel_val is None:
        return True, 0, 0

    if ppt_val is None:
        ppt_val = 0
    if excel_val is None:
        excel_val = 0

    # Handle string '-' as zero
    if ppt_val == '-':
        ppt_val = 0
    if excel_val == '-':
        excel_val = 0

    try:
        ppt_num = float(ppt_val)
        excel_num = float(excel_val)
    except (ValueError, TypeError):
        return str(ppt_val) == str(excel_val), None, None

    diff = ppt_num - excel_num

    if field_type == 'currency':
        match = abs(diff) <= TOLERANCE_CURRENCY
        diff_pct = (diff / excel_num * 100) if excel_num != 0 else (100 if diff != 0 else 0)
    elif field_type == 'percentage':
        match = abs(diff) <= TOLERANCE_PERCENTAGE
        diff_pct = (diff / excel_num * 100) if excel_num != 0 else (100 if diff != 0 else 0)
    elif field_type == 'integer':
        match = int(ppt_num) == int(excel_num)
        diff_pct = (diff / excel_num * 100) if excel_num != 0 else (100 if diff != 0 else 0)
    else:
        match = ppt_num == excel_num
        diff_pct = 0

    return match, diff, diff_pct

def main():
    # Load extracted data
    with open('ppt_extracted_data.json') as f:
        ppt_data = json.load(f)

    with open('excel_extracted_data.json') as f:
        excel_data = json.load(f)

    # Filter out total rows and build lookup dicts
    ppt_records = {make_key(r): r for r in ppt_data['records'] if not r.get('is_total')}
    excel_records = {make_key(r): r for r in excel_data['records'] if not r.get('is_total')}

    print(f"PPT records (non-total): {len(ppt_records)}")
    print(f"Excel records (non-total): {len(excel_records)}")

    # Find matches and differences
    all_keys = set(ppt_records.keys()) | set(excel_records.keys())
    ppt_only = set(ppt_records.keys()) - set(excel_records.keys())
    excel_only = set(excel_records.keys()) - set(ppt_records.keys())
    common = set(ppt_records.keys()) & set(excel_records.keys())

    print(f"\nRecords only in PPT (new): {len(ppt_only)}")
    print(f"Records only in Excel (missing from PPT): {len(excel_only)}")
    print(f"Records in both: {len(common)}")

    # Compare common records
    discrepancies = []
    matches = []

    for key in common:
        ppt_rec = ppt_records[key]
        excel_rec = excel_records[key]

        record_diffs = []
        for field, config in COMPARE_FIELDS.items():
            ppt_val = ppt_rec.get(field)
            excel_val = excel_rec.get(field)

            match, diff, diff_pct = compare_values(ppt_val, excel_val, config['type'])

            if not match:
                record_diffs.append({
                    'field': field,
                    'label': config['label'],
                    'ppt_value': ppt_val,
                    'excel_value': excel_val,
                    'difference': diff,
                    'diff_percent': diff_pct,
                    'type': config['type']
                })

        if record_diffs:
            discrepancies.append({
                'key': key,
                'market': ppt_rec.get('market'),
                'category': ppt_rec.get('category'),
                'brand': ppt_rec.get('brand'),
                'excel_row': excel_rec.get('excel_row'),
                'differences': record_diffs
            })
        else:
            matches.append(key)

    # Generate report
    report = {
        'summary': {
            'ppt_records': len(ppt_records),
            'excel_records': len(excel_records),
            'matching_records': len(matches),
            'records_with_discrepancies': len(discrepancies),
            'ppt_only_records': len(ppt_only),
            'excel_only_records': len(excel_only),
            'total_field_discrepancies': sum(len(d['differences']) for d in discrepancies)
        },
        'ppt_only': [{'key': k, 'record': ppt_records[k]} for k in sorted(ppt_only)],
        'excel_only': [{'key': k, 'record': excel_records[k]} for k in sorted(excel_only)],
        'discrepancies': discrepancies,
        'matches': list(matches)
    }

    # Save JSON report
    with open('diff_report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)

    # Print summary
    print("\n" + "="*60)
    print("DIFF REPORT SUMMARY")
    print("="*60)
    print(f"Records with discrepancies: {len(discrepancies)}")
    print(f"Total field mismatches: {report['summary']['total_field_discrepancies']}")
    print(f"Records only in PPT (need to add to Excel): {len(ppt_only)}")
    print(f"Records only in Excel (missing from PPT): {len(excel_only)}")

    # Show PPT-only records
    if ppt_only:
        print("\n--- Records in PPT but NOT in Excel (need to ADD) ---")
        for k in sorted(ppt_only):
            r = ppt_records[k]
            print(f"  {r['market']} / {r['category']} / {r['brand']}: Budget={r.get('budget_2026')}")

    # Show Excel-only records
    if excel_only:
        print("\n--- Records in Excel but NOT in PPT (FLAGGED for review) ---")
        for k in sorted(excel_only):
            r = excel_records[k]
            print(f"  {r['market']} / {r['category']} / {r['brand']}: Budget={r.get('budget_2026')} (Excel row {r.get('excel_row')})")

    # Show discrepancies by market
    print("\n--- Discrepancies by Market ---")
    by_market = defaultdict(list)
    for d in discrepancies:
        by_market[d['market']].append(d)

    for market in sorted(by_market.keys()):
        print(f"\n{market}:")
        for d in by_market[market]:
            print(f"  {d['brand']}:")
            for diff in d['differences']:
                print(f"    {diff['label']}: PPT={diff['ppt_value']} vs Excel={diff['excel_value']} (diff={diff['difference']:.2f})" if isinstance(diff['difference'], (int, float)) else f"    {diff['label']}: PPT={diff['ppt_value']} vs Excel={diff['excel_value']}")

    # Generate Excel report
    generate_excel_report(report, ppt_records, excel_records)

    print("\n" + "="*60)
    print("Reports generated:")
    print("  - diff_report.json (machine-readable)")
    print("  - diff_report.xlsx (Excel with highlighting)")
    print("="*60)

def generate_excel_report(report, ppt_records, excel_records):
    """Generate formatted Excel report"""
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Summary Sheet ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["DIFF REPORT SUMMARY", ""],
        ["", ""],
        ["PPT Records", report['summary']['ppt_records']],
        ["Excel Records", report['summary']['excel_records']],
        ["Matching Records", report['summary']['matching_records']],
        ["Records with Discrepancies", report['summary']['records_with_discrepancies']],
        ["Total Field Mismatches", report['summary']['total_field_discrepancies']],
        ["Records in PPT only (to add)", report['summary']['ppt_only_records']],
        ["Records in Excel only (review)", report['summary']['excel_only_records']],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15

    # === Discrepancies Sheet ===
    ws_disc = wb.create_sheet("Discrepancies")

    headers = ["Market", "Category", "Brand", "Excel Row", "Field", "PPT Value", "Excel Value", "Difference", "Diff %"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_disc.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    for disc in report['discrepancies']:
        for diff in disc['differences']:
            ws_disc.cell(row=row_idx, column=1, value=disc['market']).border = thin_border
            ws_disc.cell(row=row_idx, column=2, value=disc['category']).border = thin_border
            ws_disc.cell(row=row_idx, column=3, value=disc['brand']).border = thin_border
            ws_disc.cell(row=row_idx, column=4, value=disc['excel_row']).border = thin_border
            ws_disc.cell(row=row_idx, column=5, value=diff['label']).border = thin_border

            ppt_cell = ws_disc.cell(row=row_idx, column=6, value=diff['ppt_value'])
            ppt_cell.border = thin_border
            ppt_cell.fill = match_fill

            excel_cell = ws_disc.cell(row=row_idx, column=7, value=diff['excel_value'])
            excel_cell.border = thin_border
            excel_cell.fill = mismatch_fill

            diff_cell = ws_disc.cell(row=row_idx, column=8, value=diff['difference'])
            diff_cell.border = thin_border

            pct_cell = ws_disc.cell(row=row_idx, column=9, value=f"{diff['diff_percent']:.2f}%" if diff['diff_percent'] else "")
            pct_cell.border = thin_border

            row_idx += 1

    # Auto-width columns
    for col in range(1, 10):
        ws_disc.column_dimensions[get_column_letter(col)].width = 15

    # === PPT Only Sheet ===
    ws_ppt = wb.create_sheet("PPT Only (Add to Excel)")

    headers = ["Market", "Category", "Brand", "2026 Budget", "2026 Sufficient", "AWA", "CON", "PUR", "TV", "Digital", "Others", "Long Camps", "Short Camps"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_ppt.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, item in enumerate(report['ppt_only'], 2):
        r = item['record']
        ws_ppt.cell(row=row_idx, column=1, value=r.get('market'))
        ws_ppt.cell(row=row_idx, column=2, value=r.get('category'))
        ws_ppt.cell(row=row_idx, column=3, value=r.get('brand'))
        ws_ppt.cell(row=row_idx, column=4, value=r.get('budget_2026'))
        ws_ppt.cell(row=row_idx, column=5, value=r.get('sufficient_2026'))
        ws_ppt.cell(row=row_idx, column=6, value=r.get('awa'))
        ws_ppt.cell(row=row_idx, column=7, value=r.get('con'))
        ws_ppt.cell(row=row_idx, column=8, value=r.get('pur'))
        ws_ppt.cell(row=row_idx, column=9, value=r.get('tv'))
        ws_ppt.cell(row=row_idx, column=10, value=r.get('digital'))
        ws_ppt.cell(row=row_idx, column=11, value=r.get('others'))
        ws_ppt.cell(row=row_idx, column=12, value=r.get('long_campaigns'))
        ws_ppt.cell(row=row_idx, column=13, value=r.get('short_campaigns'))

    # === Excel Only Sheet ===
    ws_excel = wb.create_sheet("Excel Only (Review)")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_excel.cell(row=1, column=col_idx, value=header)
        cell.fill = warning_fill
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(report['excel_only'], 2):
        r = item['record']
        ws_excel.cell(row=row_idx, column=1, value=r.get('market'))
        ws_excel.cell(row=row_idx, column=2, value=r.get('category'))
        ws_excel.cell(row=row_idx, column=3, value=r.get('brand'))
        ws_excel.cell(row=row_idx, column=4, value=r.get('budget_2026'))
        ws_excel.cell(row=row_idx, column=5, value=r.get('sufficient_2026'))
        ws_excel.cell(row=row_idx, column=6, value=r.get('awa'))
        ws_excel.cell(row=row_idx, column=7, value=r.get('con'))
        ws_excel.cell(row=row_idx, column=8, value=r.get('pur'))
        ws_excel.cell(row=row_idx, column=9, value=r.get('tv'))
        ws_excel.cell(row=row_idx, column=10, value=r.get('digital'))
        ws_excel.cell(row=row_idx, column=11, value=r.get('others'))
        ws_excel.cell(row=row_idx, column=12, value=r.get('long_campaigns'))
        ws_excel.cell(row=row_idx, column=13, value=r.get('short_campaigns'))

    wb.save('diff_report.xlsx')

if __name__ == '__main__':
    main()
