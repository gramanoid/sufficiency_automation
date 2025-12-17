#!/usr/bin/env python3
"""
Compare Original vs Updated Excel files
Verifies that updates were actually applied correctly.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict


def normalize_value(val):
    """Normalize value for comparison"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == '-' or val == '':
            return 0
        # Check if it's a formula
        if val.startswith('='):
            return val  # Return formula as-is
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


def values_match(v1, v2, tolerance=0.01):
    """Check if two values match within tolerance"""
    if v1 == v2:
        return True

    # Both None
    if v1 is None and v2 is None:
        return True

    # One is formula, one is number
    if isinstance(v1, str) and v1.startswith('='):
        return False  # Formula vs value = different
    if isinstance(v2, str) and v2.startswith('='):
        return False

    try:
        n1 = float(v1) if v1 is not None else 0
        n2 = float(v2) if v2 is not None else 0
        return abs(n1 - n2) <= tolerance
    except (ValueError, TypeError):
        return str(v1) == str(v2)


def main():
    base_dir = Path(__file__).parent.parent

    original_path = base_dir / 'input' / 'Haleon - 2026 MEA Budget Sufficiency_271125_Final 1.xlsx'
    updated_path = base_dir / 'output' / 'Haleon - 2026 MEA Budget Sufficiency_UPDATED.xlsx'
    update_log_path = base_dir / 'output' / 'data' / 'update_log.json'

    print("=" * 70)
    print("EXCEL FILE COMPARISON: Original vs Updated")
    print("=" * 70)
    print(f"\nOriginal: {original_path.name}")
    print(f"Updated:  {updated_path.name}")

    # Load update log to know what should have changed
    with open(update_log_path) as f:
        update_log = json.load(f)

    expected_changes = update_log['changes']
    print(f"\nExpected changes from update log: {len(expected_changes)}")

    # Load both workbooks
    print("\nLoading workbooks...")
    wb_original = load_workbook(original_path, data_only=False)  # Keep formulas
    wb_updated = load_workbook(updated_path, data_only=False)

    ws_original = wb_original['2026 Sufficiency']
    ws_updated = wb_updated['2026 Sufficiency']

    # Track results
    verified_changes = []
    failed_changes = []
    unexpected_same = []

    print("\n--- Verifying Expected Changes ---\n")

    for change in expected_changes:
        row = change['row']
        col = change['col']
        field = change['field']
        old_expected = change['old_value']
        new_expected = change['new_value']
        market = change.get('market', '?')
        brand = change.get('brand', '?')

        # Get actual values
        original_val = ws_original.cell(row=row, column=col).value
        updated_val = ws_updated.cell(row=row, column=col).value

        # Normalize for comparison
        original_norm = normalize_value(original_val)
        updated_norm = normalize_value(updated_val)
        new_expected_norm = normalize_value(new_expected)

        # Check if update was applied
        update_applied = False

        # The updated value should match the new_expected value
        if values_match(updated_norm, new_expected_norm, tolerance=0.001):
            update_applied = True
            verified_changes.append({
                'row': row,
                'col': col,
                'field': field,
                'market': market,
                'brand': brand,
                'original': original_val,
                'updated': updated_val,
                'expected': new_expected
            })
        else:
            failed_changes.append({
                'row': row,
                'col': col,
                'field': field,
                'market': market,
                'brand': brand,
                'original': original_val,
                'updated': updated_val,
                'expected': new_expected,
                'issue': f"Updated value {updated_val} doesn't match expected {new_expected}"
            })

    # Summary
    print(f"Verified changes: {len(verified_changes)}/{len(expected_changes)}")
    print(f"Failed changes: {len(failed_changes)}")

    if failed_changes:
        print("\n--- FAILED CHANGES (need investigation) ---")
        for fc in failed_changes[:20]:
            print(f"  Row {fc['row']}, Col {fc['col']} ({fc['field']}):")
            print(f"    Market/Brand: {fc['market']}/{fc['brand']}")
            print(f"    Original: {fc['original']}")
            print(f"    Updated:  {fc['updated']}")
            print(f"    Expected: {fc['expected']}")
            print(f"    Issue: {fc['issue']}")
        if len(failed_changes) > 20:
            print(f"  ... and {len(failed_changes) - 20} more")

    # Sample of verified changes
    print("\n--- Sample Verified Changes ---")
    for vc in verified_changes[:10]:
        original_display = vc['original']
        if isinstance(original_display, str) and len(original_display) > 30:
            original_display = original_display[:30] + "..."
        print(f"  Row {vc['row']}, {vc['field']}: {original_display} -> {vc['updated']}")
        print(f"    ({vc['market']}/{vc['brand']})")

    # Check for formula -> value conversions
    formula_to_value = [vc for vc in verified_changes
                        if isinstance(vc['original'], str) and vc['original'].startswith('=')]

    print(f"\n--- Formula to Value Conversions: {len(formula_to_value)} ---")
    if formula_to_value:
        print("  (These cells had formulas that were replaced with PPT values)")
        for ftv in formula_to_value[:5]:
            print(f"  Row {ftv['row']}, {ftv['field']}: Formula -> {ftv['updated']}")

    # Additional check: Compare a few random cells that should NOT have changed
    print("\n--- Spot Check: Cells That Should Be Unchanged ---")
    unchanged_checks = [
        (5, 4, 'Budget 2026 (E5)'),  # Budget shouldn't change
        (5, 3, 'Brand (D5)'),  # Brand name
        (5, 10, 'AWA (J5)'),  # Should match if no change needed
    ]

    for row, col, desc in unchanged_checks:
        orig = ws_original.cell(row=row, column=col).value
        upd = ws_updated.cell(row=row, column=col).value

        # Check if this cell was in the changes list
        was_changed = any(c['row'] == row and c['col'] == col for c in expected_changes)

        if was_changed:
            print(f"  {desc}: Changed (as expected)")
        elif values_match(orig, upd):
            print(f"  {desc}: Unchanged (correct)")
        else:
            print(f"  {desc}: UNEXPECTEDLY CHANGED! {orig} -> {upd}")

    # Final verdict
    print("\n" + "=" * 70)
    if len(failed_changes) == 0:
        print("VERIFICATION RESULT: SUCCESS")
        print(f"All {len(verified_changes)} expected changes were correctly applied.")
        print(f"{len(formula_to_value)} formula cells were converted to values (from PPT).")
    else:
        print("VERIFICATION RESULT: ISSUES FOUND")
        print(f"{len(failed_changes)} changes did not apply correctly.")
    print("=" * 70)

    # Save detailed report
    report = {
        'summary': {
            'expected_changes': len(expected_changes),
            'verified_changes': len(verified_changes),
            'failed_changes': len(failed_changes),
            'formula_to_value': len(formula_to_value)
        },
        'verified_changes': verified_changes,
        'failed_changes': failed_changes,
        'formula_conversions': formula_to_value
    }

    report_path = base_dir / 'output' / 'reports' / 'excel_comparison_report.json'
    with open(report_path, 'w') as f:
        json.dump(report, f, indent=2, default=str)

    print(f"\nDetailed report saved to: {report_path}")

    return len(failed_changes) == 0


if __name__ == '__main__':
    import sys
    success = main()
    sys.exit(0 if success else 1)
