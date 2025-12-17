#!/usr/bin/env python3
"""
Apply PPT Data to Excel - Updates Excel to match PPT values
Preserves ALL formatting (fonts, colors, borders, column widths, merged cells)
Only changes cell VALUES, not styles
"""

import json
from openpyxl import load_workbook
from datetime import datetime

# Column mapping for "2026 Sufficiency" sheet (1-based Excel columns)
FIELD_TO_COL = {
    'budget_2026': 5,      # E
    'sufficient_2026': 7,  # G
    'gap_gbp': 8,          # H
    'gap_pct': 9,          # I
    'awa': 10,             # J
    'con': 11,             # K
    'pur': 12,             # L
    'tv': 18,              # R
    'digital': 19,         # S
    'others': 22,          # V
    'long_campaigns': 30,  # AD
    'short_campaigns': 32, # AF
    'long_pct': 34,        # AH
}

def normalize_str(s):
    """Normalize string for comparison"""
    if not s:
        return ''
    return str(s).strip().upper().replace(' ', '').replace('-', '')

def values_differ(ppt_val, excel_val, field_type):
    """Check if values are different enough to warrant update"""
    if ppt_val is None:
        return False

    if ppt_val == '-':
        ppt_val = 0
    if excel_val == '-':
        excel_val = 0
    if excel_val is None:
        excel_val = 0

    try:
        ppt_num = float(ppt_val)
        excel_num = float(excel_val) if excel_val else 0
    except (ValueError, TypeError):
        return str(ppt_val) != str(excel_val)

    if field_type == 'currency':
        return abs(ppt_num - excel_num) > 1.0
    elif field_type == 'percentage':
        return abs(ppt_num - excel_num) > 0.001
    elif field_type == 'integer':
        return int(ppt_num) != int(excel_num)
    else:
        return ppt_num != excel_num

def main():
    xlsx_path = 'Haleon - 2026 MEA Budget Sufficiency_271125_Final 1.xlsx'

    with open('ppt_extracted_data.json') as f:
        ppt_data = json.load(f)

    with open('excel_extracted_data.json') as f:
        excel_data = json.load(f)

    # Build lookup from (market, category, brand) -> excel_row
    excel_row_lookup = {}
    for rec in excel_data['records']:
        key = (
            normalize_str(rec.get('market')),
            normalize_str(rec.get('category')),
            normalize_str(rec.get('brand'))
        )
        excel_row_lookup[key] = rec.get('excel_row')

    print(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    ws = wb['2026 Sufficiency']

    changes_log = []
    updates_made = 0
    records_not_found = []

    for ppt_rec in ppt_data['records']:
        if ppt_rec.get('is_total'):
            continue

        market = ppt_rec.get('market')
        category = ppt_rec.get('category')
        brand = ppt_rec.get('brand')

        key = (normalize_str(market), normalize_str(category), normalize_str(brand))
        excel_row = excel_row_lookup.get(key)

        if not excel_row:
            records_not_found.append(f"{market} / {category} / {brand}")
            continue

        for field, col in FIELD_TO_COL.items():
            ppt_val = ppt_rec.get(field)

            if ppt_val is None:
                continue

            current_val = ws.cell(row=excel_row, column=col).value

            if field in ['budget_2026', 'sufficient_2026', 'gap_gbp']:
                field_type = 'currency'
            elif field in ['long_campaigns', 'short_campaigns']:
                field_type = 'integer'
            else:
                field_type = 'percentage'

            if values_differ(ppt_val, current_val, field_type):
                cell = ws.cell(row=excel_row, column=col)
                old_value = cell.value

                if field_type == 'integer':
                    new_value = int(ppt_val) if ppt_val not in ['-', None] else 0
                elif ppt_val == '-':
                    new_value = 0
                else:
                    new_value = float(ppt_val)

                cell.value = new_value
                updates_made += 1

                changes_log.append({
                    'row': excel_row,
                    'col': col,
                    'field': field,
                    'market': market,
                    'category': category,
                    'brand': brand,
                    'old_value': old_value,
                    'new_value': new_value
                })

    output_path = 'Haleon - 2026 MEA Budget Sufficiency_UPDATED.xlsx'
    wb.save(output_path)

    log_output = {
        'timestamp': datetime.now().isoformat(),
        'source_file': xlsx_path,
        'output_file': output_path,
        'summary': {
            'total_updates': updates_made,
            'records_not_found': len(records_not_found)
        },
        'changes': changes_log,
        'records_not_found': records_not_found
    }

    with open('update_log.json', 'w') as f:
        json.dump(log_output, f, indent=2, default=str)

    print("\n" + "="*60)
    print("UPDATE COMPLETE")
    print("="*60)
    print(f"Total cell updates: {updates_made}")
    print(f"PPT records not found in Excel: {len(records_not_found)}")

    if records_not_found:
        print("\nRecords in PPT but not found in Excel:")
        for r in records_not_found:
            print(f"  - {r}")

    print(f"\nOutput saved to: {output_path}")
    print(f"Change log saved to: update_log.json")

    print("\n--- Sample Changes Made ---")
    for change in changes_log[:20]:
        print(f"  Row {change['row']}, {change['field']}: {change['old_value']} -> {change['new_value']}")
        print(f"    ({change['market']} / {change['brand']})")

    if len(changes_log) > 20:
        print(f"  ... and {len(changes_log) - 20} more changes")

if __name__ == '__main__':
    main()
